import hashlib
import re
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore


def make_doc_id(title, fecha_p):
    return hashlib.sha1(f"{title}_{fecha_p}".encode()).hexdigest()

def extract_filename(disposition, content_type, url, opt_title):
    if disposition:
        match = re.search(r'filename="?([^"]+)"?', disposition)
        if match:
            filename = match.group(1)
            # Extract extension from filename
            ext = "." + filename.split(".")[-1] if "." in filename else ""
            return {"filename": filename.split(".")[0], "extension": ext}

    if "rtf" in content_type.lower():
        ext = ".rtf"
    elif "pdf" in content_type.lower():
        ext = ".pdf"
    elif "word" in content_type.lower() or "officedocument" in content_type.lower():
        ext = ".docx"
    else:
        ext = ""

    return {"filename": url.split("/")[-1] or opt_title, "extension": ext}


def generate_excel_report(path: str, rows: list[dict], title: str = "Inventario") -> None:
    """Generate an Excel report from a list of row dictionaries.
    
    Creates a workbook with:
    - Title and generation timestamp
    - Table with columns: Descargado, Fuente, Título
    """
    
    if Workbook is None:
        raise ImportError("openpyxl is required for Excel report generation. Install it with: pip install openpyxl")
    
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Add title
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    
    # Add generation timestamp
    ws['A2'] = f"Generado: {created_at}"
    ws['A2'].font = Font(size=10)
    
    # Add headers
    headers = ["Descargado", "Fuente", "Título"]
    header_row = 4
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    data_row = header_row + 1
    for r in rows:
        ws.cell(row=data_row, column=1).value = str(r.get("downloaded_at", ""))[:20]
        ws.cell(row=data_row, column=2).value = str(r.get("source", ""))[:20]
        ws.cell(row=data_row, column=3).value = str(r.get("title", ""))
        data_row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    
    # Save workbook
    wb.save(path)


def get_asp_data(soup):
    viewstate = soup.select_one("input[name='__VIEWSTATE']")["value"]    
    eventvalidation = soup.select_one("input[name='__EVENTVALIDATION']")["value"]
    viewstate_generator = soup.select_one("input[name='__VIEWSTATEGENERATOR']")["value"] # Datos asp.net

    return {"__VIEWSTATE":viewstate, "__EVENTVALIDATION":eventvalidation, "__VIEWSTATEGENERATOR":viewstate_generator}

def parse_ajax_response(response_text):
    # El formato AJAX de ASP.NET es: longitud|tipo|id|contenido|
    # Queremos el contenido que pertenece al updatePanel
    partes = response_text.split('|')
    
    # Buscamos el ViewState nuevo que SIEMPRE cambia en el POST
    new_asp_data = {}
    for i in range(len(partes)):
        if partes[i] in ["__VIEWSTATE", "__EVENTVALIDATION", "__VIEWSTATEGENERATOR"]:
            new_asp_data[partes[i]] = partes[i+1]
            
    # Intentamos extraer el HTML del bloque 'updatePanel'
    html_content = ""
    if "updatePanel" in partes:
        idx = partes.index("updatePanel")
        html_content = partes[idx + 2] # El HTML suele estar 2 posiciones después
        
    return html_content, new_asp_data